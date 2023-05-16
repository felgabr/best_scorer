# ==================================================================================
# PACKAGES
# ==================================================================================
# Standard
import argparse
import datetime
import json
import logging.config
import os
import socket
import subprocess
import sys
import time
import warnings

# Third Party
import numpy as np
import pyspark.sql.functions as f
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, numbers
# ==================================================================================
# SETTINGS
# ==================================================================================
warnings.filterwarnings('ignore')
# ==================================================================================
# PARAMETERS pt1
# ==================================================================================
parser = argparse.ArgumentParser()
parser.add_argument('-c', '--config', type=str, default='config.json')
parser.add_argument('-l', '--log', type=str, default='logging.conf')
args = parser.parse_args()
# GLOBAL
abs_path = os.getcwd()
user = os.environ['USER']
# Load config file
conf_filepath = os.path.join(abs_path, 'config.json')
__CONFIG__ = json.load(open(conf_filepath))
# General
project = __CONFIG__['project']
app_name = f'{project}_merge_prfmce'
schema = __CONFIG__['schema']
# ==================================================================================
# LOGGER
# ==================================================================================
log_filepath = os.path.join(abs_path, 'logging.conf')
logging.config.fileConfig(fname=log_filepath)
logger = logging.getLogger(project)
log_file, stdlog_file = __CONFIG__['logger_files']
# ==================================================================================
# LOCAL
# ==================================================================================
from supplementary import DateAux, LakeAux, OpenPyXlManager, PersistManager
from utils import build_spark_session, send_email
# ==================================================================================
# PARAMETERS pt2
# ==================================================================================
# Tables
master_src = __CONFIG__['table_outputs']['score']
# Mailing
dest = __CONFIG__['from']
recp = __CONFIG__['to']
recp_cp = __CONFIG__['cc']
# Output
project = __CONFIG__['project']
alias = __CONFIG__['alias']
file_merge = __CONFIG__['rep_perfnce']
# ML config
config = __CONFIG__['conf']
# ==================================================================================
# MAIN
# ==================================================================================
try:
    logger.info('Running')
    logger.info(f'Hostname: {socket.gethostname()}')

    start_time = time.time()
    # ==============================================================================
    # SPARK SESSION
    # ==============================================================================
    spark = build_spark_session(app_name, 
                                 __CONFIG__['size'], 
                                 args=[('spark.rpc.message.maxSize', 256)], 
                                 jars=[], 
                                 py_files=[__CONFIG__['py_file0'],
                                           __CONFIG__['py_file1']],
                                 path_prefix='log/')
    spark.conf.set("spark.sql.legacy.allowCreatingManagedTableUsingNonemptyLocation", "true")

    sc = spark.sparkContext
    
    logger.info('Spark applicationId: ' + str(sc.applicationId))
    logger.info(f'Spark Parameters: \n{spark.sparkContext.getConf().getAll()}')
    # ==============================================================================
    # PROCESS
    # ==============================================================================
    logger.info('Start reference data')

    master_src = LakeAux.get_most_recent_table(spark, schema, master_src)
    logger.info(f'Master: {master_src}')
    df = spark.read.table(f'{schema}.{master_src}')
    
    logger.info('Start load reports')
    reports = []
    reports.append(PersistManager.read_csv_from_hdfs(spark, __CONFIG__['outputs']['metrics']).toPandas())
    reports.append(PersistManager.read_csv_from_hdfs(spark, __CONFIG__['outputs']['feat_import']).toPandas())
    reports.append(PersistManager.read_csv_from_hdfs(spark, __CONFIG__['outputs']['ks_train']).toPandas())
    reports.append(PersistManager.read_csv_from_hdfs(spark, __CONFIG__['outputs']['ks_test']).toPandas())
    row_dates = df.filter(f.col(config['spine']['ind_split_col']) == 'val') \
                  .select(config['spine']['date_col']).orderBy(f.col(config['spine']['date_col']))\
                                                      .distinct().collect()
    
    sheet_names = ['METRICS', 'FEAT_IMPORT_MODEL', 'KS_TRAIN', 'KS_TEST']
    for i, date in enumerate(row_dates):
        reports.append(PersistManager.read_csv_from_hdfs(spark, __CONFIG__['outputs']['ks_val']+f'_{i}').toPandas())
        date = ''.join(date[0].split('-')[:2])
        sheet_names.append(f'KS_VAL{date}')

    logger.info('Start Build merged')
    # Defining formats
    # Font and fill color
    font = Font(name='Arial')
    fontbold = Font(name='Arial', bold=True)
    header_font = Font(name='Arial', color='FFFFFF', bold=True)
    fill_header = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    cols_names = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                  'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                  'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']
    
    # Cols to format
    columns_m = cols_names[:reports[0].shape[1]]
    perc_cols_m = cols_names[1:reports[0].shape[1]]

    columns_fi = cols_names[:reports[1].shape[1]]
    perc_cols_fi = cols_names[1:reports[1].shape[1]]

    columns_ks = cols_names[:reports[2].shape[1]]
    perc_cols_ks = ['H', 'I', 'O']
    
    # Sheet creation
    wb = Workbook()
    for name in sheet_names:
        ws = wb.create_sheet(name)
    wb.remove(wb['Sheet'])
    
    # Writing data
    for i, sheet in enumerate(wb.worksheets):
        OpenPyXlManager.write_data(sheet, reports[i])

    for i, sheet in enumerate(wb.worksheets):
        if i == 0:
            # General format
            OpenPyXlManager.format_cols(sheet, columns_m, 'Pandas', font, fontbold)
            # Percentage format
            OpenPyXlManager.perc_format_cols(sheet, perc_cols_m)
        elif i == 1:
            OpenPyXlManager.format_cols(sheet, columns_fi, 'Pandas', font, fontbold)
            OpenPyXlManager.perc_format_cols(sheet, perc_cols_fi)
        else:
            OpenPyXlManager.format_cols(sheet, columns_ks, 'Pandas', font, fontbold)
            OpenPyXlManager.perc_format_cols(sheet, perc_cols_ks)             
                    
    # Header
    for sheet in wb.worksheets:
        OpenPyXlManager.header_format(sheet, header_font, fill_header)
    
    # Change columns width
    for i, sheet in enumerate(wb.worksheets):
        for col in cols_names:
            if col == 'A':
                if i == 0:
                    sheet.column_dimensions[col].width = 20
                elif i == 1:
                    sheet.column_dimensions[col].width = 50
                else:
                    sheet.column_dimensions[col].width = 7
            else:
                sheet.column_dimensions[col].width = 15
                
    logger.info('Start Saving')
    wb.save(file_merge)
    wb.close()
    
    logger.info(PersistManager.copy_from_local_to_hdfs(file_merge, project))
# ==================================================================================
# MAILING
# ==================================================================================
    logger.info('Start send email')
    end_time = time.time()
    subject = 'RUN: ' + app_name + ' - OK'
    message = f'''Dear coallegues,<br> Time elapsed merging reports
                  was {DateAux.time_elapsed(end_time, start_time)} hours.
                  <br> <br> Best Regards.''' 
    send_email(sender=dest, to=recp, cc=recp_cp, message=message, subject=subject,
                 attch=[log_file, stdlog_file, file_merge])   

except Exception as e:
    logger.error('------------------------------------------------\n')
    logger.error('EXECUTION ERROR\n')
    logger.error('------------------------------------------------\n')
    logger.error('ERROR: {}'.format(str(e)))
    end_time = time.time()
    
    subject = 'RUN: ' + app_name + ' - KO'
    message = f'''Dear coallegues,<br> Time elapsed in this attempt to merge reports
                  was {DateAux.time_elapsed(end_time, start_time)} hours.
                  <br> <br> Best Regards.''' 
    send_email(sender=dest, to=recp, cc=recp_cp, message=message, subject=subject,
                 attch=[log_file, stdlog_file]